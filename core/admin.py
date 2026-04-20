from django.contrib import admin
from .models import Publication, Personne, Delocalisation, DelocalisationPlantation, Contact, Audience, Newsletter

#================ PUBLICATIONS =================
@admin.register(Publication)
class PublicationAdmin(admin.ModelAdmin):
    list_display = ('titre', 'date_publication')
    search_fields = ('titre',)


# ================= PERSONNE =================
@admin.register(Personne)
class PersonneAdmin(admin.ModelAdmin):
    list_display = ('nom', 'postnom', 'prenom', 'numero_identite', 'telephone')
    search_fields = ('nom', 'postnom', 'numero_identite', 'telephone')


# ================= DELOCALISATION =================
@admin.register(Delocalisation)
class DelocalisationAdmin(admin.ModelAdmin):
    change_form_template = "admin/delocalisation_change_form.html"
    list_display = (
        'personne',
        'type_propriete',
        'type_delocalisation',
        'ancienne_adresse',
        'nouvelle_adresse',
        'montant_compensation',
        'date_demande'
    )

    list_filter = ('type_delocalisation', 'type_propriete', 'date_demande')

    search_fields = (
        'personne__nom',
        'personne__postnom',
        'personne__numero_identite',
        'ancienne_adresse',
        'nouvelle_adresse'
    )

    autocomplete_fields = ['personne']

    readonly_fields = ('date_demande',)
    fieldsets = (
        ("Informations générales", {
            "fields": ("personne", "type_propriete", "type_delocalisation")
        }),

        ("Ancienne adresse", {
            "fields": ("ancienne_adresse", "ancienne_latitude", "ancienne_longitude")
        }),

        ("Nouvelle adresse", {
            "fields": ("nouvelle_adresse", "nouvelle_latitude", "nouvelle_longitude")
        }),

        ("Compensation", {
            "fields": ("montant_compensation",)
        }),

        ("Meta", {
            "fields": ("date_demande",)
        }),
    )


# ================= DELOCALISATION DES PLANTATIONS =================
from django.http import HttpResponse
from django.urls import path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import datetime


@admin.register(DelocalisationPlantation)
class DelocalisationPlantationAdmin(admin.ModelAdmin):
    change_form_template = "admin/cpd_plantation_change_form.html"
    change_list_template = "admin/cpd_plantation_changelist.html"

    list_display = (
        'nom', 'postnom', 'prenom',
        'type_piece_identite', 'site_adresse',
        'nature_culture', 'type_structure_maison',
        'date_enregistrement'
    )

    list_filter = ('type_piece_identite', 'type_structure_maison', 'date_enregistrement')

    search_fields = (
        'nom', 'postnom', 'prenom',
        'numero_carte_electeur', 'numero_passeport',
        'numero_permis_conduire', 'numero_attestation_naissance',
        'nature_culture', 'site_adresse'
    )

    readonly_fields = ('date_enregistrement',)

    fieldsets = (
        ("Identité (obligatoire)", {
            "fields": ("nom", "postnom", "prenom", "photo", "site_adresse")
        }),
        ("Pièce d'identité (obligatoire)", {
            "description": "Choisissez le type de pièce, puis renseignez le numéro correspondant.",
            "fields": (
                "type_piece_identite",
                "numero_carte_electeur",
                "numero_passeport",
                "numero_permis_conduire",
                "numero_attestation_naissance",
            )
        }),
        ("Champ (facultatif)", {
            "fields": ("superficie", "nature_culture"),
            "classes": ("collapse",),
        }),
        ("Structure maison (facultatif)", {
            "fields": ("superficie_maison", "type_structure_maison"),
            "classes": ("collapse",),
        }),
        ("Méta", {
            "fields": ("date_enregistrement",)
        }),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-excel/', self.admin_site.admin_view(self.export_excel_view), name='cpd_export_excel'),
        ]
        return custom_urls + urls

    def export_excel_view(self, request):
        """
        Exporte les résultats filtrés en Excel.
        Reprend les mêmes filtres que la liste admin affichée.
        """
        # Récupérer le queryset filtré (en respectant les filtres admin)
        from django.contrib.admin.views.main import ChangeList
        try:
            cl = ChangeList(
                request, self.model, self.list_display,
                self.list_display_links, self.list_filter, self.date_hierarchy,
                self.search_fields, self.list_select_related, self.list_per_page,
                self.list_max_show_all, self.list_editable, self, self.sortable_by,
                self.search_help_text,
            )
            queryset = cl.get_queryset(request)
        except Exception:
            queryset = self.get_queryset(request)

        # Créer le workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Liste CPD"

        # ===== STYLES =====
        title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1a5276', end_color='1a5276', fill_type='solid')
        cell_font = Font(name='Calibri', size=10)
        cell_font_bold = Font(name='Calibri', size=10, bold=True)
        thin_border = Border(
            left=Side(style='thin', color='B0B0B0'),
            right=Side(style='thin', color='B0B0B0'),
            top=Side(style='thin', color='B0B0B0'),
            bottom=Side(style='thin', color='B0B0B0'),
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        even_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

        # ===== TITRE DU DOCUMENT =====
        headers = [
            'N°', 'Nom', 'Post-nom', 'Prénom', 'Site / Adresse',
            "Type pièce d'identité",
            'Superficie du champs (m²)', 'Nature de culture',
            'Superficie parcelle (m²)', 'Type structure maison',
            "Date d'enregistrement"
        ]
        nb_cols = len(headers)

        # Ligne de titre fusionnée
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
        title_cell = ws.cell(row=1, column=1,
                             value="CPD — Commission Provinciale de Délocalisation")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 35

        # Sous-titre avec date et nombre
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nb_cols)
        subtitle_cell = ws.cell(
            row=2, column=1,
            value=f"Liste exportée le {datetime.datetime.now().strftime('%d/%m/%Y à %H:%M')} — {queryset.count()} individu(s)"
        )
        subtitle_cell.font = Font(name='Calibri', size=10, italic=True, color='555555')
        subtitle_cell.alignment = center_align
        ws.row_dimensions[2].height = 22

        # ===== EN-TÊTES =====
        header_row = 4
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 28

        # ===== DONNÉES =====
        TYPE_PIECE_DISPLAY = dict(DelocalisationPlantation.TYPE_PIECE_CHOICES)
        TYPE_STRUCTURE_DISPLAY = dict(DelocalisationPlantation.TYPE_STRUCTURE_CHOICES)

        for idx, obj in enumerate(queryset, 1):
            row_num = header_row + idx
            row_data = [
                idx,
                obj.nom,
                obj.postnom,
                obj.prenom,
                obj.site_adresse or '',
                TYPE_PIECE_DISPLAY.get(obj.type_piece_identite, obj.type_piece_identite or ''),
                float(obj.superficie) if obj.superficie else '',
                obj.nature_culture or '',
                float(obj.superficie_maison) if obj.superficie_maison else '',
                TYPE_STRUCTURE_DISPLAY.get(obj.type_structure_maison, obj.type_structure_maison or ''),
                obj.date_enregistrement.strftime('%d/%m/%Y %H:%M') if obj.date_enregistrement else '',
            ]

            row_fill = even_fill if idx % 2 == 0 else None

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = cell_font_bold if col_idx == 1 else cell_font
                cell.alignment = center_align if col_idx in (1, 6, 7, 9, 10, 11) else left_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

        # ===== LARGEURS COLONNES =====
        col_widths = [6, 18, 18, 18, 25, 24, 22, 22, 22, 22, 22]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # ===== LIGNE TOTAL =====
        total_row = header_row + queryset.count() + 1
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        total_cell = ws.cell(row=total_row, column=1, value=f"TOTAL : {queryset.count()} individu(s)")
        total_cell.font = Font(name='Calibri', size=11, bold=True, color='003366')
        total_cell.alignment = center_align
        for col_idx in range(1, nb_cols + 1):
            ws.cell(row=total_row, column=col_idx).border = Border(
                top=Side(style='medium', color='003366'),
                bottom=Side(style='medium', color='003366'),
            )

        # ===== FIGER VOLETS =====
        ws.freeze_panes = f'A{header_row + 1}'

        # ===== RÉPONSE HTTP =====
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"CPD_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response



# ================= CONTACT =================
@admin.register(Contact)
class ContactAdmin(admin.ModelAdmin):
    list_display = ('nom_complet', 'email', 'date_envoi', 'lu')
    list_filter = ('lu', 'date_envoi')
    search_fields = ('nom_complet', 'email', 'message')
    readonly_fields = ('nom_complet', 'email', 'message', 'date_envoi')
    list_editable = ('lu',)

    def has_add_permission(self, request):
        return False


# ================= AUDIENCE =================
@admin.register(Audience)
class AudienceAdmin(admin.ModelAdmin):
    list_display = ('nom_complet', 'poste', 'statut', 'date_demande')
    list_filter = ('statut', 'date_demande')
    search_fields = ('nom_complet', 'poste')
    readonly_fields = ('nom_complet', 'poste', 'photo_identite', 'lettre_demande', 'date_demande')
    list_editable = ('statut',)

    def has_add_permission(self, request):
        return False

# ================= NEWSLETTER =================
@admin.register(Newsletter)
class NewsletterAdmin(admin.ModelAdmin):
    list_display = ('email', 'date_inscription', 'actif')
    list_filter = ('actif', 'date_inscription')
    search_fields = ('email',)
    list_editable = ('actif',)
    readonly_fields = ('date_inscription',)
